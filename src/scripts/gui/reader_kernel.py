from os.path import splitext
from tkinter import Tk, messagebox

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .reader_gui import ReaderGUI
import csv

class ReaderKernel(object):
    def __init__(self, parent, file_name):
        self.parent = parent
        self.file_name = file_name
        excel_formats = [".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"]
        _, file_extension = splitext(file_name)
        file_format = "csv"
        if file_extension in excel_formats:
            file_format = "excel"
        self.root = Tk()
        self.app = ReaderGUI(self.root, self, file_format)
        self.root.mainloop()
        self.root.destroy()

    def cmd_get_sheets(self):
        try:
            wb2 = load_workbook(self.file_name)
            return wb2.get_sheet_names()
        except InvalidFileException:
            return None

    def cmd_ok(self):
        if self.app.file_format.get() == "csv":
            data = self.load_csv()
        else:
            data = self.load_excel()
        if data:
            try:
                header = data[int(self.app.header.get())-1]
                data[0:int(self.app.header.get())] = []
            except ValueError:
                header = data.pop(0)
            
            # Cast data to ints, on error skip row.
            tmp_data = []
            for row in data:
                try:
                    tmp_data.append([int(j) for j in row])
                except ValueError:
                    continue
                except TypeError:
                    continue
            data = tmp_data
            sex = []
            age = []
            ffan = {}
            cols_to_del = []
            # Iterate over the headers and search for the values given in age_symb and sex_symb.
            try:
                for i, head in reversed(list(enumerate(header))):
                    if head == self.app.age_symb.get():
                        age = self.get_col(data, i)
                        self.del_col(data, i)
                        cols_to_del.append(i)
                    elif head == self.app.sex_symb.get():
                        sex = self.get_col(data, i)
                        self.del_col(data, i)
                        cols_to_del.append(i)
                    elif head.startswith(self.app.ffan_init.get()):
                        ffan[head.replace(self.app.ffan_init.get(), "", 1)] = self.get_col(data, i)
                        self.del_col(data, i)
                        cols_to_del.append(i)
            except AttributeError:
                # Ends up here if the file is wrongly formated
                messagebox.showerror("Import error",
                                     "Unable to import data due to wrongly formated file.")
                self.root.quit()
                return

            header = [i for j, i in enumerate(header) if j not in cols_to_del]
            ffan["All"] = [1]*len(self.get_col(data, 0))

            self.parent.data = data
            self.parent.header = header
            self.parent.age = age
            self.parent.sex = sex
            self.parent.ffan = ffan

        self.root.quit()
    
    def load_csv(self):
        with open(self.file_name) as csvfile:
            data = list(csv.reader(csvfile, delimiter=self.app.delimiter.get()))
        return data

    def load_excel(self):
        # Read in the workbook.
        workbook = load_workbook(self.file_name)
        worksheet = workbook.get_sheet_by_name(self.app.sheet.get())
        data = [[cell.value for cell in row] for row in worksheet.rows]
        return data

    def valid_col(self, column):
        column = list(set(column))
        for value in range(1, 6):
            try:
                column.remove(value)
            except ValueError:
                pass
        if column:
            return False
        return True

    def del_col(self, data, col_idx):
        for row in data:
            del row[col_idx]

    def get_col(self, data, col_idx):
        return [row[col_idx] for row in data]

    def cmd_cancel(self):
        self.root.quit()