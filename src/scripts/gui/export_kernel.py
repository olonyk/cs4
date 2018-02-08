from tkinter import Tk

from openpyxl import Workbook
from openpyxl.chart import AreaChart, Reference, Series
from openpyxl.compat import range
from openpyxl.styles import Alignment, Font
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

from .export_gui import ExportGUI

class ExportKernel(object):
    def __init__(self, parent, result_table, file_name):
        self.parent = parent
        self.result_table = result_table
        self.file_name = file_name
        self.root = Tk()
        self.app = ExportGUI(self.root, self)
        self.root.mainloop()
        self.root.destroy()
    
    def cmd_cancel(self):
        self.root.destroy()

    def cmd_export(self):
        bkg_color = self.app.colors["color_3"].get().replace("#", "")
        bkg_fill = PatternFill(patternType='solid', fgColor=Color(rgb=bkg_color))

        head_color = self.app.colors["color_1"].get().replace("#", "")
        head_fill = PatternFill(patternType='solid', fgColor=Color(rgb=head_color))

        head_font_color = self.app.colors["font_color_1"].get().replace("#", "")
        head_font = Font(bold=True, color=head_font_color)

        body_font_color = self.app.colors["font_color_2"].get().replace("#", "")
        body_font = Font(bold=False, color=body_font_color)

        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        row_offset = 1
        col_offset = 1
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.title = "CS4_output"

        for xl_row in range(1, 100):
            for xl_col in range(1, 100):
                active_cell = work_sheet.cell(column=xl_col, row=xl_row)
                active_cell.fill = bkg_fill

        for xl_row, res_row in enumerate(self.result_table):
            for xl_col, res_col in enumerate(res_row):
                active_cell = work_sheet.cell(column=xl_col+col_offset,
                                              row=xl_row+row_offset,
                                              value="{0}".format(res_col))
                # Left column header
                if not xl_col:
                    active_cell.fill = head_fill
                    active_cell.font = head_font
                    active_cell.alignment = align_right
                # Top row header
                elif not xl_row:
                    active_cell.fill = head_fill
                    active_cell.font = head_font
                    active_cell.alignment = align_center
                # Data row at the bottom
                elif xl_row >= len(self.result_table)-3:
                    active_cell.value = int(res_col)
                # Style names
                else:
                    active_cell.font = body_font
                    active_cell.alignment = align_left
        
        # Create chart
        values = Reference(work_sheet,
                           min_col=col_offset+1,
                           max_col=len(self.result_table[0])+col_offset-1,
                           min_row=len(self.result_table)+row_offset-1,
                           max_row=len(self.result_table)+row_offset-1)
        chart = AreaChart()
        chart.add_data(values, from_rows=True)
        chart.legend = None
        chart.title = "Core Sound analysis"

        # set a pattern for the whole series
        series = chart.series[0]
        fill = PatternFillProperties()
        fill.foreground = ColorChoice(prstClr="red")
        fill.background = ColorChoice(prstClr="blue")
        series.graphicalProperties.pattFill = fill

        work_sheet.add_chart(chart, "E15")

        work_book.save(filename=self.file_name)
        self.root.destroy()
